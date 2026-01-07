from fastapi import APIRouter, Depends, HTTPException, Request, Response, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
import uuid
from uuid import UUID
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import case, func
from app.core.auth import get_db, get_current_admin
from app.models.invitation import Invitation, InvitationStatus
from app.models.invitation_response import InvitationResponse
from app.models.template import InvitationTemplate
from app.models.rsvp import RsvpStatus
from app.schemas.invitation import InvitationAdminListItem, InvitationCreate, InvitationOut
from app.schemas.invitation_response import InvitationResponseCreate, InvitationResponseOut
from app.schemas.invitation_import import InvitationImportResult, InvitationImportCreatedItem, InvitationImportErrorItem
from slugify import slugify
from datetime import datetime
import os

import openpyxl

router = APIRouter(prefix="/api/invitations", tags=["invitations"])


@router.get("/export", dependencies=[Depends(get_current_admin)])
def export_invitations(request: Request, db: Session = Depends(get_db)):
    invitations = db.query(Invitation).order_by(Invitation.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invitations"

    origin = request.headers.get("origin")
    base_url = (os.getenv("PUBLIC_INVITE_BASE_URL") or origin or "").rstrip("/")

    headers = [
        "recipient_salutation",
        "recipient_name",
        "recipient_title",
        "status",
        "slug",
        "invite_url",
        "rsvp_status",
        "attendee_count",
        "created_at",
    ]
    ws.append(headers)

    for inv in invitations:
        slug = inv.slug or ""
        invite_path = f"/invite/{slug}" if slug else ""
        invite_url = f"{base_url}{invite_path}" if base_url and slug else invite_path

        ws.append(
            [
                inv.recipient_salutation or "",
                inv.recipient_name,
                inv.recipient_title,
                inv.status.value if hasattr(inv.status, "value") else str(inv.status),
                slug,
                invite_url,
                inv.rsvp_status.value if hasattr(inv.rsvp_status, "value") else str(inv.rsvp_status),
                int(inv.attendee_count or 0),
                inv.created_at.isoformat() if inv.created_at else "",
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"invitations_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _generate_unique_slug(db: Session, base_text: str, used: set[str]) -> str:
    base = slugify(base_text)
    candidate = base
    suffix = 2
    while candidate in used or db.query(Invitation).filter(Invitation.slug == candidate).first():
        candidate = f"{base}-{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _normalize_header(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = s.replace("\u00a0", " ")
    s = " ".join(s.split())
    s = s.replace("-", "_")
    s = s.replace(" ", "_")
    return s


def _cell_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v).strip()


def _safe_cell_text(row: tuple[object, ...], idx: int | None) -> str:
    if idx is None:
        return ""
    if idx < 0:
        return ""
    if idx >= len(row):
        return ""
    return _cell_text(row[idx])


@router.get("/", dependencies=[Depends(get_current_admin)])
def list_invitations(db: Session = Depends(get_db)) -> list[InvitationAdminListItem]:
    rows = (
        db.query(
            Invitation,
            func.count(InvitationResponse.id).label("responses"),
            func.coalesce(
                func.sum(case((InvitationResponse.response == RsvpStatus.ATTENDING, 1), else_=0)),
                0,
            ).label("attending"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            InvitationResponse.response == RsvpStatus.ATTENDING,
                            func.coalesce(InvitationResponse.attendee_count, 1),
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("attending_people"),
            func.coalesce(
                func.sum(case((InvitationResponse.response == RsvpStatus.DECLINED, 1), else_=0)),
                0,
            ).label("declined"),
        )
        .outerjoin(InvitationResponse, InvitationResponse.invitation_id == Invitation.id)
        .group_by(Invitation.id)
        .order_by(Invitation.created_at.desc())
        .all()
    )
    result: list[InvitationAdminListItem] = []
    for inv, responses, attending, attending_people, declined in rows:
        base = InvitationOut.model_validate(inv).model_dump()
        result.append(
            InvitationAdminListItem(
                **base,
                responses=int(responses),
                attending=int(attending),
                attending_people=int(attending_people),
                declined=int(declined),
            )
        )
    return result

@router.post("/", response_model=InvitationOut, dependencies=[Depends(get_current_admin)])
def create_invitation(invitation: InvitationCreate, db: Session = Depends(get_db)):
    payload = invitation.dict()

    # Default salutation if not provided (older clients / optional field).
    salutation = payload.get("recipient_salutation")
    if salutation is None or (isinstance(salutation, str) and not salutation.strip()):
        payload["recipient_salutation"] = "Ông"

    # Default draft if status not provided.
    status_value = payload.pop("status", None)
    if status_value:
        try:
            payload["status"] = InvitationStatus(status_value)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status")

    db_invitation = Invitation(**payload)

    # Rule: slug only generated when publishing.
    if db_invitation.status == InvitationStatus.published and not db_invitation.slug:
        base = slugify(f"{db_invitation.title}-{db_invitation.recipient_name}")
        candidate = base
        suffix = 2
        while db.query(Invitation).filter(Invitation.slug == candidate).first():
            candidate = f"{base}-{suffix}"
            suffix += 1
        db_invitation.slug = candidate

    db.add(db_invitation)
    db.commit()
    db.refresh(db_invitation)
    return db_invitation


@router.post("/import", response_model=InvitationImportResult, dependencies=[Depends(get_current_admin)])
async def import_invitations(
    template_id: UUID = Form(...),
    status_value: str = Form("draft"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # Validate status
    try:
        status_enum = InvitationStatus(status_value)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid status")

    # Validate template
    tpl = db.query(InvitationTemplate).filter(InvitationTemplate.id == template_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    if not tpl.event_time or not tpl.event_location:
        raise HTTPException(status_code=400, detail="Template must have event_time and event_location")

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid .xlsx file")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Empty sheet")

    headers = [_normalize_header(h) for h in header_row]

    # Allow a few common header variants.
    aliases = {
        "recipient_salutation": {
            "recipient_salutation",
            "salutation",
            "xung_ho",
            "xungho",
            "ong_ba",
            "ông_bà",
            "ong",
            "ba",
        },
        "recipient_name": {
            "recipient_name",
            "name",
            "ten",
            "ho_ten",
            "họ_tên",
            "nguoi_nhan",
            "người_nhận",
        },
        "recipient_title": {
            "recipient_title",
            "title",
            "chuc_danh",
            "chức_danh",
            "chuc_vu",
            "chức_vụ",
        },
    }

    index_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        for canonical, keys in aliases.items():
            if h in keys and canonical not in index_map:
                index_map[canonical] = i

    if "recipient_name" not in index_map or "recipient_title" not in index_map:
        raise HTTPException(
            status_code=400,
            detail="Missing required columns: recipient_name and recipient_title",
        )

    created_items: list[InvitationImportCreatedItem] = []
    errors: list[InvitationImportErrorItem] = []
    used_slugs: set[str] = set()
    created = 0
    skipped = 0

    # Excel rows are 1-based; header is row 1, data starts at row 2.
    excel_row_num = 1
    for row in rows_iter:
        excel_row_num += 1

        name = _safe_cell_text(row, index_map.get("recipient_name"))
        title = _safe_cell_text(row, index_map.get("recipient_title"))
        salutation = _safe_cell_text(row, index_map.get("recipient_salutation"))

        if not name and not title and not salutation:
            skipped += 1
            continue

        if not name:
            errors.append(InvitationImportErrorItem(row=excel_row_num, message="Missing recipient_name"))
            continue
        if not title:
            errors.append(InvitationImportErrorItem(row=excel_row_num, message="Missing recipient_title"))
            continue

        inv = Invitation(
            title=tpl.title,
            company_name=tpl.company_name,
            recipient_salutation=salutation.strip() or "Ông",
            recipient_name=name,
            recipient_title=title,
            content=tpl.content,
            event_time=tpl.event_time,
            event_location=tpl.event_location,
            google_map_url=tpl.google_map_url,
            schedule=tpl.schedule,
            status=status_enum,
        )

        if inv.status == InvitationStatus.published and not inv.slug:
            inv.slug = _generate_unique_slug(db, f"{inv.title}-{inv.recipient_name}", used_slugs)

        db.add(inv)
        db.flush()  # populate id
        created += 1
        created_items.append(InvitationImportCreatedItem(row=excel_row_num, id=inv.id, slug=inv.slug))

    db.commit()

    return InvitationImportResult(created=created, skipped=skipped, items=created_items, errors=errors)

@router.get("/{slug}", response_model=InvitationOut)
def get_invitation(slug: str, db: Session = Depends(get_db)):
    invitation = db.query(Invitation).filter(Invitation.slug == slug, Invitation.status == InvitationStatus.published).first()
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")
    return invitation

@router.post("/{slug}/response", response_model=InvitationResponseOut)
def respond_invitation(
    slug: str,
    payload: InvitationResponseCreate,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
):
    invitation = db.query(Invitation).filter(Invitation.slug == slug, Invitation.status == InvitationStatus.published).first()
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")

    response_value = payload.response
    if response_value == RsvpStatus.PENDING:
        raise HTTPException(status_code=400, detail="Invalid response")

    attendee_count = payload.attendee_count
    if response_value == RsvpStatus.DECLINED:
        attendee_count = 0
    else:
        if attendee_count is None:
            attendee_count = 1
        if attendee_count < 1 or attendee_count > 20:
            raise HTTPException(status_code=400, detail="attendee_count must be between 1 and 20")

    # 1 invitation has exactly 1 response (last write wins).
    db_response = (
        db.query(InvitationResponse)
        .filter(InvitationResponse.invitation_id == invitation.id)
        .order_by(InvitationResponse.created_at.desc())
        .first()
    )

    if db_response:
        db_response.response = response_value
        db_response.attendee_count = attendee_count
    else:
        db_response = InvitationResponse(
            invitation_id=invitation.id,
            responder_id=invitation.id,
            response=response_value,
            attendee_count=attendee_count,
        )
        db.add(db_response)

    invitation.rsvp_status = response_value
    invitation.attendee_count = attendee_count if response_value == RsvpStatus.ATTENDING else 0
    db.commit()
    db.refresh(db_response)

    one_year = 60 * 60 * 24 * 365
    response.set_cookie(
        f"invite_choice_{invitation.id}",
        response_value.value,
        httponly=False,
        samesite="lax",
        max_age=one_year,
    )
    return db_response


@router.delete("/{invitation_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(get_current_admin)])
def delete_invitation(invitation_id: UUID, db: Session = Depends(get_db)):
    invitation = db.query(Invitation).filter(Invitation.id == invitation_id).first()
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")

    db.query(InvitationResponse).filter(InvitationResponse.invitation_id == invitation_id).delete(synchronize_session=False)
    db.delete(invitation)
    db.commit()
    return None
